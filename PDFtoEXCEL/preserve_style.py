import openpyxl
import os

def copy_cell_styles(source_cell, target_cell):
    if source_cell.has_style:
        target_cell.font = openpyxl.styles.Font(
            name=source_cell.font.name,
            size=source_cell.font.size,
            bold=source_cell.font.bold,
            italic=source_cell.font.italic,
            vertAlign=source_cell.font.vertAlign,
            underline=source_cell.font.underline,
            strike=source_cell.font.strike,
            color=source_cell.font.color
        )
        target_cell.border = openpyxl.styles.Border(
            left=source_cell.border.left,
            right=source_cell.border.right,
            top=source_cell.border.top,
            bottom=source_cell.border.bottom,
            diagonal=source_cell.border.diagonal,
            diagonal_direction=source_cell.border.diagonal_direction,
            outline=source_cell.border.outline,
            vertical=source_cell.border.vertical,
            horizontal=source_cell.border.horizontal
        )
        target_cell.fill = openpyxl.styles.PatternFill(
            fill_type=source_cell.fill.fill_type,
            start_color=source_cell.fill.start_color,
            end_color=source_cell.fill.end_color
        )
        target_cell.number_format = source_cell.number_format
        target_cell.protection = openpyxl.styles.Protection(
            locked=source_cell.protection.locked,
            hidden=source_cell.protection.hidden
        )
        target_cell.alignment = openpyxl.styles.Alignment(
            horizontal=source_cell.alignment.horizontal,
            vertical=source_cell.alignment.vertical,
            text_rotation=source_cell.alignment.text_rotation,
            wrap_text=source_cell.alignment.wrap_text,
            shrink_to_fit=source_cell.alignment.shrink_to_fit,
            indent=source_cell.alignment.indent
        )

def combine_workbooks(directory, output_file):
    combined_wb = openpyxl.Workbook()
    combined_ws = combined_wb.active
    combined_ws.title = "Combined Data"

    current_row = 1

    for i in range(1, 5769):
        filename = f"report_page{i}.xlsx"
        filepath = os.path.join(directory, filename)

        if os.path.exists(filepath):
            print(f"Processing {filename}")
            wb = openpyxl.load_workbook(filepath)
            ws = wb.active

            for row in ws.iter_rows():
                for cell in row:
                    combined_cell = combined_ws.cell(row=current_row, column=cell.column, value=cell.value)
                    copy_cell_styles(cell, combined_cell)
                current_row += 1

        else:
            print(f"File {filename} not found")

    combined_wb.save(output_file)
    print(f"Combined workbook saved as {output_file}")

# Specify the directory containing the workbooks and the name of the output file
directory = "C:\\Users\\Ubaid Rehman\\Desktop\\PDF2EXCEL\\PDF2EXCEL\\Report_Excel\\"
output_file = "preserved_combined_workbook.xlsx"

combine_workbooks(directory, output_file)
