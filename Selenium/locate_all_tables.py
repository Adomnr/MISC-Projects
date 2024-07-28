from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from bs4 import BeautifulSoup
import pandas as pd
from openpyxl import load_workbook, Workbook
import os

# Initialize the Chrome driver
driver = webdriver.Chrome()

try:
    # Open the specified URL
    driver.get('https://us.soccerway.com/teams/comparison/?competition_ids%5B%5D=263&team_ids%5B%5D=203&competition_ids%5B%5D=953&team_ids%5B%5D=113')

    # Wait for the tables to load
    WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.TAG_NAME, 'table')))

    # Get the page source
    page_source = driver.page_source

    # Use BeautifulSoup to parse the page source
    soup = BeautifulSoup(page_source, 'html.parser')

    # Find all tables in the page
    tables = soup.find_all('table')

    # Initialize a list to store the desired rows
    desired_rows = []

    # Iterate over tables to find the one starting with "General statistics"
    for table in tables:
        # Check if the table starts with "General statistics"
        first_row = table.find('tr')
        if first_row:
            first_cell = first_row.find('th')
            if first_cell and 'General statistics' in first_cell.text.strip():
                # Extract table rows
                rows = table.find_all('tr')
                for row in rows:
                    cells = row.find_all(['td', 'th'])
                    cell_texts = [cell.text.strip() for cell in cells]
                    if cell_texts:
                        # Filter for desired rows
                        if cell_texts[0] in [
                            'Matches played', 'Clean sheets',
                            'Avg. goals scored p/m', 'Avg. goals conceded p/m',
                            'Failed to score'
                        ]:
                            desired_rows.append(cell_texts)
                break  # Exit loop after finding the desired table

finally:
    # Close the driver
    driver.quit()

# Convert the extracted data into a pandas DataFrame
df = pd.DataFrame(desired_rows)

# Function to convert strings to numbers where possible
def convert_to_number(value):
    try:
        if '.' in value:
            return float(value)
        else:
            return int(value)
    except ValueError:
        return value

# Apply the conversion function to the DataFrame
df = df.applymap(convert_to_number)

# Define the file path
file_path = 'data.xlsx'

# Check if the file exists and is a valid Excel file
if not os.path.exists(file_path):
    # Create a new Excel file
    wb = Workbook()
    wb.save(file_path)

try:
    # Load the existing Excel file
    book = load_workbook(file_path)
except Exception as e:
    print(f"Error loading the Excel file: {e}")
    # Create a new valid Excel file if loading fails
    wb = Workbook()
    wb.save(file_path)
    book = load_workbook(file_path)

# Define the start row and column for the data
startrow = 3  # 4th row (0-indexed)
startcol = 2  # 3rd column (0-indexed)

# Get the active sheet
sheet = book.active

# Write the DataFrame to the sheet
for r_idx, row in enumerate(df.values, start=startrow + 1):
    for c_idx, value in enumerate(row, start=startcol + 1):
        cell = sheet.cell(row=r_idx, column=c_idx)
        cell.value = value

# Save the workbook
book.save(file_path)
