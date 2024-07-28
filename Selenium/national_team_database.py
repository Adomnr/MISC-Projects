from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import Select
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
import os
import time
import openpyxl

# Bypass SSL verification
os.environ['WDM_SSL_VERIFY'] = '0'

# Initialize the WebDriver
driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()))

# Open the target URL
url = "https://us.soccerway.com/teams/comparison/?team_ids%5B%5D=16210&team_ids%5B%5D=16210"
driver.get(url)

# Wait for the page to load completely
driver.implicitly_wait(10)


# Function to extract options from a dropdown
def get_dropdown_options(dropdown):
    select = Select(dropdown)
    options = [option.text for option in select.options]
    return options


# Function to select an option and extract data from dependent dropdowns
def extract_data_for_selection(dropdown_index, option):
    dropdown = Select(driver.find_elements(By.CSS_SELECTOR, "select")[dropdown_index])
    dropdown.select_by_visible_text(option)
    time.sleep(0)  # Wait for the page to update

    # Re-find the dependent dropdown elements
    dependent_dropdowns = driver.find_elements(By.CSS_SELECTOR, "select")

    # Extract data from the dependent dropdowns
    dependent_data = {}
    dependent_data['Dropdown 8'] = get_dropdown_options(dependent_dropdowns[7])
    dependent_data['Dropdown 9'] = get_dropdown_options(dependent_dropdowns[8])
    # dependent_data['Dropdown 11'] = get_dropdown_options(dependent_dropdowns[10])
    # dependent_data['Dropdown 12'] = get_dropdown_options(dependent_dropdowns[11])

    return dependent_data


# Create a new Excel workbook and select the active sheet
wb = openpyxl.Workbook()
ws = wb.active

# Initialize row counter
row_counter = 1

# Locate the initial dropdown elements
dropdown7 = Select(driver.find_elements(By.CSS_SELECTOR, "select")[7])

dropdown7_options = [option.text for option in dropdown7.options if option.text]
print(dropdown7_options)
# Iterate over each option in Dropdown 7 and extract data
for option in dropdown7_options:
    print(f"Selecting {option} in Dropdown 8")
    all_data = extract_data_for_selection(7, option)

    # Store the current selected country in row 1
    ws.cell(row=row_counter, column=1, value=option)

    # Move to the next row
    row_counter += 1
    time.sleep(0.2)

    dropdown8 = Select(driver.find_elements(By.CSS_SELECTOR, "select")[8])
    dropdown8_options = [option2.text for option2 in dropdown8.options if option2.text]
    print(dropdown8_options)
    for idx, option3 in enumerate(dropdown8_options):
        ws.cell(row=row_counter, column=idx + 1, value=option3)
    row_counter += 3

# Save the workbook
wb.save("dropdown_options_national.xlsx")

# Close the WebDriver
driver.quit()
