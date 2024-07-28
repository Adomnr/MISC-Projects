import openpyxl
import pandas as pd
import tkinter as tk
from tkinter import ttk

import os
import time
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import Select, WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.chrome.options import Options


# List of countries
countries = [
    "Afghanistan","Albania", "Algeria","American Samoa", "Andorra", "Angola", "Antigua and Barbuda", "Argentina", "Armenia", "Aruba",
    "Australia", "Austria", "Azerbaijan", "Bahrain", "Bangladesh", "Barbados", "Belarus", "Belgium",
    "Belize", "Benin", "Bermuda", "Bhutan", "Bolivia", "Bosnia and Herzegovina", "Botswana", "Brazil",
    "British Virgin Islands", "Brunei Darussalam", "Bulgaria", "Burkina Faso", "Burundi", "Cambodia",
    "Cameroon", "Canada", "Cape Verde", "Cayman Islands", "Chile", "China PR", "Chinese Taipei",
    "Colombia", "Congo DR", "Congo", "Cook Islands", "Costa Rica", "Croatia", "Cuba", "Curaçao",
    "Cyprus", "Czechia", "Côte d'Ivoire", "Denmark", "Djibouti", "Dominican Republic", "Ecuador",
    "Egypt", "El Salvador", "England", "Estonia", "Eswatini", "Ethiopia", "Faroe Islands", "Fiji",
    "Finland", "France", "French Guiana", "Gabon", "Gambia", "Georgia", "Germany", "Ghana",
    "Gibraltar", "Greece", "Grenada", "Guadeloupe", "Guam", "Guatemala", "Guinea", "Guyana", "Haiti",
    "Honduras", "Hong Kong, China", "Hungary", "Iceland", "India", "Indonesia", "Iran", "Iraq",
    "Israel", "Italy", "Jamaica", "Japan", "Jordan", "Kazakhstan", "Kenya", "Korea Republic", "Kosovo",
    "Kuwait", "Kyrgyz Republic", "Laos", "Latvia", "Lebanon", "Lesotho", "Liberia", "Libya",
    "Lithuania", "Luxembourg", "Macao", "Madagascar", "Malawi", "Malaysia", "Maldives", "Mali",
    "Malta", "Martinique", "Mauritania", "Mauritius", "Mexico", "Moldova", "Mongolia", "Montenegro",
    "Morocco", "Mozambique", "Myanmar", "Nepal", "Netherlands", "New Caledonia", "New Zealand",
    "Nicaragua", "Nigeria", "North Macedonia", "Northern Ireland", "Norway", "Oman", "Pakistan",
    "Palestine", "Panama", "Papua New Guinea", "Paraguay", "Peru", "Philippines", "Poland", "Portugal",
    "Puerto Rico", "Qatar", "Republic of Ireland", "Romania", "Russia", "Rwanda", "Réunion",
    "San Marino", "Saudi Arabia", "Scotland", "Senegal", "Serbia", "Sierra Leone", "Singapore",
    "Slovakia", "Slovenia", "Solomon Islands", "Somalia", "South Africa", "Spain", "Sri Lanka",
    "St. Kitts and Nevis", "Sudan", "Suriname", "Sweden", "Switzerland", "Syria", "São Tomé e Príncipe",
    "Tahiti", "Tajikistan", "Tanzania", "Thailand", "Togo", "Trinidad and Tobago", "Tunisia",
    "Turkmenistan", "Turks and Caicos Islands", "Türkiye", "USA", "Uganda", "Ukraine",
    "United Arab Emirates", "Uruguay", "Uzbekistan", "Venezuela", "Vietnam", "Wales", "Yemen",
    "Zambia", "Zimbabwe"
]

file_path_club = 'dropdown_options_filtered_club.xlsx'  # Update this with the correct file path for club
file_path_national = 'dropdown_options_national.xlsx'  # Update this with the correct file path for national

# Function to read and process Excel data for club
def read_excel_data(file_path):
    df = pd.read_excel(file_path, header=None)
    df.dropna(how='all', inplace=True)
    data_dict = {}
    current_country = None
    is_competition_row = False
    for index, row in df.iterrows():
        row_data = [str(cell) for cell in row if pd.notna(cell)]
        if row[0] in countries:
            current_country = row[0]
            is_competition_row = True
            data_dict[current_country] = {'competitions': [], 'teams': {}}
        elif current_country:
            if is_competition_row:
                competition = row[0]
                data_dict[current_country]['competitions'].append(competition)
                data_dict[current_country]['teams'][competition] = []
                is_competition_row = False
            else:
                data_dict[current_country]['teams'][competition].extend(row_data)
                is_competition_row = True
    return data_dict

# Function to read and process Excel data for national
def read_excel_data_national(file_path):
    df = pd.read_excel(file_path, header=None)
    df.dropna(how='all', inplace=True)
    data_dict = {}
    current_country = None
    for index, row in df.iterrows():
        row_data = [str(cell) for cell in row if pd.notna(cell)]
        if row[0] in countries:
            current_country = row[0]
            data_dict[current_country] = []
        elif current_country:
            data_dict[current_country].extend(row_data)
    return data_dict

# Read club and national data
data_dict_club = read_excel_data(file_path_club)
data_dict_national = read_excel_data_national(file_path_national)

# GUI Creation
def update_competitions(event, dropdown_box, dropdown_comp, data_dict):
    selected_country = dropdown_box.get()
    competitions = data_dict[selected_country]['competitions']
    dropdown_comp['values'] = competitions

def update_teams(event, dropdown_comp, dropdown_team, country_dropdown, data_dict):
    selected_competition = dropdown_comp.get()
    selected_country = country_dropdown.get()
    teams = data_dict[selected_country]['teams'][selected_competition]
    dropdown_team['values'] = teams

def update_ui(event):
    selected_type = dropdown_1.get()
    if selected_type == "National":
        dropdown_6.grid_remove()
        dropdown_7.grid_remove()
        dropdown_2.set('')
        dropdown_3.set('')
        dropdown_4.set('')
        dropdown_5.set('')
        dropdown_2.bind("<<ComboboxSelected>>", lambda event: update_teams_national(event, dropdown_2, dropdown_4, data_dict_national))
        dropdown_3.bind("<<ComboboxSelected>>", lambda event: update_teams_national(event, dropdown_3, dropdown_5, data_dict_national))
    else:
        dropdown_6.grid()
        dropdown_7.grid()
        dropdown_2.set('')
        dropdown_3.set('')
        dropdown_4.set('')
        dropdown_5.set('')
        dropdown_2.bind("<<ComboboxSelected>>", lambda event: update_competitions(event, dropdown_2, dropdown_4, data_dict_club))
        dropdown_3.bind("<<ComboboxSelected>>", lambda event: update_competitions(event, dropdown_3, dropdown_5, data_dict_club))
        dropdown_4.bind("<<ComboboxSelected>>", lambda event: update_teams(event, dropdown_4, dropdown_6, dropdown_2, data_dict_club))
        dropdown_5.bind("<<ComboboxSelected>>", lambda event: update_teams(event, dropdown_5, dropdown_7, dropdown_3, data_dict_club))

def update_teams_national(event, country_dropdown, dropdown_comp, data_dict):
    selected_country = country_dropdown.get()
    if selected_country in data_dict:
        teams = data_dict[selected_country]  # Get the row under the country
        dropdown_comp['values'] = teams
    else:
        dropdown_comp['values'] = []

def compare_two_teams_club(selection, country1,competition1, team1, country2, competition2, team2):
    # Bypass SSL verification
    os.environ['WDM_SSL_VERIFY'] = '0'

    chrome_options = Options()
    chrome_options.add_experimental_option("detach", True)  # Prevents Chrome from closing immediately

    # Initialize the WebDriver
    driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()))

    # Open the target URL
    url = "https://us.soccerway.com/matches/2024/07/23/korea-republic/k-league/seongnam-ilhwa/chunnam-dragons/4308862/head2head/"
    driver.get(url)

    # Wait for the page to load completely
    driver.implicitly_wait(4)

    # Function to extract options from a dropdown
    def get_dropdown_options(dropdown):
        select = Select(dropdown)
        options = [option.text for option in select.options]
        return options

    # Function to select an option and extract data from dependent dropdowns
    def extract_data_for_selection(dropdown_index, option):
        dropdown = Select(driver.find_elements(By.CSS_SELECTOR, "select")[dropdown_index])
        dropdown.select_by_visible_text(option)
        time.sleep(2)  # Wait for the page to update

        # Re-find the dependent dropdown elements
        dependent_dropdowns = driver.find_elements(By.CSS_SELECTOR, "select")

        # Extract data from the dependent dropdowns
        dependent_data = {}
        dependent_data['Dropdown 8'] = get_dropdown_options(dependent_dropdowns[7])
        dependent_data['Dropdown 9'] = get_dropdown_options(dependent_dropdowns[8])
        dependent_data['Dropdown 11'] = get_dropdown_options(dependent_dropdowns[10])
        dependent_data['Dropdown 12'] = get_dropdown_options(dependent_dropdowns[11])

        return dependent_data

    # Locate the initial dropdown elements
    dropdown7 = Select(driver.find_elements(By.CSS_SELECTOR, "select")[6])
    # dropdown10 = Select(driver.find_elements(By.CSS_SELECTOR, "select")[9])

    dropdown7_options = [option.text for option in dropdown7.options if option.text]
    # dropdown10_options = [option.text for option in dropdown10.options if option.text]

    # Dictionary to store the extracted data
    all_data = {}

    # for option in dropdown7_options:
    #     print(f"Selecting {option} in Dropdown 7")
    #     all_data[f"Dropdown 7 - {option}"] = extract_data_for_selection(6, option)

    # Iterate over each option in Dropdown 10 and extract data
    # for option in dropdown10_options:
    #     print(f"Selecting {option} in Dropdown 10")
    #     all_data[f"Dropdown 10 - {option}"] = extract_data_for_selection(9, option)

    def compare_two_teams_function(selection, country1, competition1, team1, country2, competition2, team2):
        if selection == "Club":
            print(dropdown7_options)
            for option in dropdown7_options:
                print(f"Selecting {option} in Dropdown 7")
                if country1 == option:
                    all_data[f"Dropdown 7 - {option}"] = extract_data_for_selection(6, option)
                if option == country1:
                    break
            time.sleep(1)
            dropdown8 = Select(driver.find_elements(By.CSS_SELECTOR, "select")[7])
            dropdown8_options = [option2.text for option2 in dropdown8.options if option2.text]
            print(dropdown8_options)

            for option in dropdown8_options:
                print(f"Selecting {option} in Dropdown 8")
                if competition1 == option:
                    all_data[f"Dropdown 8 - {option}"] = extract_data_for_selection(7, option)
                if option == competition1:
                    break
            time.sleep(1)
            dropdown9 = Select(driver.find_elements(By.CSS_SELECTOR, "select")[8])
            dropdown9_options = [option.text for option in dropdown9.options if option.text]
            for option in dropdown9_options:
                print(f"Selecting {option} in Dropdown 9")
                if team1 == option:
                    all_data[f"Dropdown 9 - {option}"] = extract_data_for_selection(8, option)
                if option == team1:
                    break
            dropdown10 = Select(driver.find_elements(By.CSS_SELECTOR, "select")[9])
            dropdown10_options = [option.text for option in dropdown10.options if option.text]
            for option in dropdown10_options:
                print(f"Selecting {option} in Dropdown 10")
                if country2 == option:
                    all_data[f"Dropdown 10 - {option}"] = extract_data_for_selection(9, option)
                if option == country2:
                    break
            dropdown11 = Select(driver.find_elements(By.CSS_SELECTOR, "select")[10])
            dropdown11_options = [option.text for option in dropdown11.options if option.text]
            for option in dropdown11_options:
                print(f"Selecting {option} in Dropdown 11")
                if competition2 == option:
                    all_data[f"Dropdown 11 - {option}"] = extract_data_for_selection(10, option)
                if option == competition2:
                    break
            dropdown12 = Select(driver.find_elements(By.CSS_SELECTOR, "select")[11])
            dropdown12_options = [option.text for option in dropdown12.options if option.text]
            for option in dropdown12_options:
                print(f"Selecting {option} in Dropdown 12")
                if team2 == option:
                    all_data[f"Dropdown 12 - {option}"] = extract_data_for_selection(11, option)
                if option == team2:
                    break
            inputs = driver.find_elements(By.XPATH, "//input[@type='button' or @type='submit']")

            # Look for the specific button with text "Compare two teams" and click it
            for input_elem in inputs:
                input_text = input_elem.get_attribute("value")
                if input_text and "compare two teams" in input_text.lower():
                    print(f"Found 'Compare two teams' button: {input_text}")
                    input_elem.click()  # Click the button
                    break

            # Extract all table elements
            tables = driver.find_elements(By.TAG_NAME, "table")

            # Print the number of tables found
            print(f"Number of tables found: {len(tables)}")

            # Prepare data storage
            extracted_data = []

            # Check if the 15th table exists
            if len(tables) >= 15:
                # Extract data from the 15th table
                table = tables[14]  # Index 14 for the 15th table
                print("\nTable 15:")
                rows = table.find_elements(By.CSS_SELECTOR, "tr")

                # Specify the row indices we are interested in (0-based index)
                row_indices = [3, 10, 11, 12, 15]  # Row 4, 11, 12, 13, and 16

                for index in row_indices:
                    if index < len(rows):
                        row = rows[index]
                        data = row.find_elements(By.CSS_SELECTOR, "td")
                        data_text = [d.text for d in data]
                        extracted_data.append(data_text)
                        print(f"Row {index + 1}: {data_text}")
                    else:
                        print(f"Row {index + 1} does not exist.")
            else:
                print("The 15th table does not exist on this page.")

            # Close the driver
            driver.quit()

            # Load an existing workbook
            workbook_path = "extracted_data.xlsx"  # Replace with your actual workbook path
            workbook = openpyxl.load_workbook(workbook_path)
            sheet = workbook.active  # Replace with your actual sheet name if needed

            # Write the extracted data to the Excel sheet starting from row 2 and column 2
            start_row = 4
            start_col = 4

            for i, data in enumerate(extracted_data):
                for j, value in enumerate(data):
                    try:
                        # Try converting to integer first
                        numeric_value = int(value)
                    except ValueError:
                        try:
                            # If integer conversion fails, try converting to float
                            numeric_value = float(value.replace('m', ''))  # Handle minute notation
                        except ValueError:
                            # If both conversions fail, keep it as a string
                            numeric_value = value
                    cell = sheet.cell(row=start_row + i, column=start_col + j)
                    cell.value = numeric_value  # Only change the value, retain the existing format

            # Save the updated workbook
            workbook.save(workbook_path)
            print(f"Data has been updated in {workbook_path}")

        else:
            pass

    compare_two_teams_function(selection, country1, competition1, team1, country2, competition2, team2)

def compare_two_teams_national(selection, country1, competition1, team1, country2, competition2, team2):
    # Bypass SSL verification
    os.environ['WDM_SSL_VERIFY'] = '0'

    chrome_options = Options()
    chrome_options.add_experimental_option("detach", True)  # Prevents Chrome from closing immediately

    # Initialize the WebDriver
    driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()))

    # Open the target URL
    url = "https://us.soccerway.com/teams/comparison/?team_ids%5B%5D=14495&team_ids%5B%5D=16210"
    driver.get(url)

    # Wait for the page to load completely
    driver.implicitly_wait(4)

    # Function to extract options from a dropdown
    def get_dropdown_options(dropdown):
        select = Select(dropdown)
        options = [option.text for option in select.options]
        return options

    # Function to select an option and extract data from dependent dropdowns
    def extract_data_for_selection(dropdown_index, option):
        dropdown = Select(driver.find_elements(By.CSS_SELECTOR, "select")[dropdown_index])
        dropdown.select_by_visible_text(option)
        time.sleep(2)  # Wait for the page to update

        # Re-find the dependent dropdown elements
        dependent_dropdowns = driver.find_elements(By.CSS_SELECTOR, "select")

        # Extract data from the dependent dropdowns
        dependent_data = {}

        return dependent_data

    # Locate the initial dropdown elements
    dropdown6 = Select(driver.find_elements(By.CSS_SELECTOR, "select")[5])

    dropdown6_options = [option.text for option in dropdown6.options if option.text]
    # dropdown10_options = [option.text for option in dropdown10.options if option.text]

    # Dictionary to store the extracted data
    all_data = {}


    def compare_two_teams_function(selection, country1, competition1, team1, country2, competition2, team2):
        if selection == "National":
            print(dropdown6_options)
            for option in dropdown6_options:
                print(f"Selecting {option} in Dropdown 6")
                if country1 == option:
                    all_data[f"Dropdown 6 - {option}"] = extract_data_for_selection(5, option)
                if option == country1:
                    break
            time.sleep(1)
            dropdown7 = Select(driver.find_elements(By.CSS_SELECTOR, "select")[6])
            dropdown7_options = [option2.text for option2 in dropdown7.options if option2.text]
            print(dropdown7_options)

            for option in dropdown7_options:
                print(f"Selecting {option} in Dropdown 7")
                if competition1 == option:
                    all_data[f"Dropdown 7 - {option}"] = extract_data_for_selection(6, option)
                if option == competition1:
                    break
            time.sleep(1)
            dropdown8 = Select(driver.find_elements(By.CSS_SELECTOR, "select")[7])
            dropdown8_options = [option.text for option in dropdown8.options if option.text]
            for option in dropdown8_options:
                print(f"Selecting {option} in Dropdown 8")
                if country2 == option:
                    all_data[f"Dropdown 8 - {option}"] = extract_data_for_selection(7, option)
                if option == country2:
                    break
            dropdown9 = Select(driver.find_elements(By.CSS_SELECTOR, "select")[8])
            dropdown9_options = [option.text for option in dropdown9.options if option.text]
            for option in dropdown9_options:
                print(f"Selecting {option} in Dropdown 9")
                if competition2 == option:
                    all_data[f"Dropdown 9 - {option}"] = extract_data_for_selection(8, option)
                if option == competition2:
                    break
            inputs = driver.find_elements(By.XPATH, "//input[@type='button' or @type='submit']")

            # Look for the specific button with text "Compare two teams" and click it
            for input_elem in inputs:
                input_text = input_elem.get_attribute("value")
                if input_text and "compare two teams" in input_text.lower():
                    print(f"Found 'Compare two teams' button: {input_text}")
                    input_elem.click()  # Click the button
                    break
            # Extract all table elements
            tables = driver.find_elements(By.TAG_NAME, "table")

            # Print the number of tables found
            print(f"Number of tables found: {len(tables)}")

            # Prepare data storage
            extracted_data = []

            # Check if the 15th table exists
            if len(tables) >= 15:
                # Extract data from the 15th table
                table = tables[14]  # Index 14 for the 15th table
                print("\nTable 15:")
                rows = table.find_elements(By.CSS_SELECTOR, "tr")

                # Specify the row indices we are interested in (0-based index)
                row_indices = [3, 10, 11, 12, 15]  # Row 4, 11, 12, 13, and 16

                for index in row_indices:
                    if index < len(rows):
                        row = rows[index]
                        data = row.find_elements(By.CSS_SELECTOR, "td")
                        data_text = [d.text for d in data]
                        extracted_data.append(data_text)
                        print(f"Row {index + 1}: {data_text}")
                    else:
                        print(f"Row {index + 1} does not exist.")
            else:
                print("The 15th table does not exist on this page.")

            # Close the driver
            driver.quit()

            # Load an existing workbook
            workbook_path = "extracted_data.xlsx"  # Replace with your actual workbook path
            workbook = openpyxl.load_workbook(workbook_path)
            sheet = workbook.active  # Replace with your actual sheet name if needed

            # Write the extracted data to the Excel sheet starting from row 2 and column 2
            start_row = 4
            start_col = 4

            for i, data in enumerate(extracted_data):
                for j, value in enumerate(data):
                    try:
                        # Try converting to integer first
                        numeric_value = int(value)
                    except ValueError:
                        try:
                            # If integer conversion fails, try converting to float
                            numeric_value = float(value.replace('m', ''))  # Handle minute notation
                        except ValueError:
                            # If both conversions fail, keep it as a string
                            numeric_value = value
                    cell = sheet.cell(row=start_row + i, column=start_col + j)
                    cell.value = numeric_value  # Only change the value, retain the existing format

            # Save the updated workbook
            workbook.save(workbook_path)
            print(f"Data has been updated in {workbook_path}")

        else:
            pass

    compare_two_teams_function(selection, country1, competition1, team1, country2, competition2, team2)
    driver.quit()
def compare_two_teams_data(*args):
    selection = dropdown_1.get()
    country1 = dropdown_2.get()
    competition1 = dropdown_4.get()
    team1 = dropdown_6.get()
    country2 = dropdown_3.get()
    competition2 = dropdown_5.get()
    team2 = dropdown_7.get()
    print(selection,"     ", country1,"     ",team1,"     ",country2,"     ",team2)
    if team1 == NameError or team2 ==NameError:
        team1 = ""
        team2 = ""
    if selection == "Club":
        compare_two_teams_club(selection,country1,competition1,team1,country2,competition2,team2)
    else:
        compare_two_teams_national(selection,country1,competition1,team1,country2,competition2,team2)



# Create the main window
root = tk.Tk()
root.title("Football Teams Comparison")

# Dropdown Box 1
label_1 = tk.Label(root, text="Select Type")
label_1.grid(row=0, column=1)
dropdown_1 = ttk.Combobox(root, values=["Club", "National"])
dropdown_1.grid(row=1, column=1)
dropdown_1.bind("<<ComboboxSelected>>", update_ui)

# Dropdown Box 2
label_2 = tk.Label(root, text="Country 1")
label_2.grid(row=2, column=0)
dropdown_2 = ttk.Combobox(root, values=countries)
dropdown_2.grid(row=3, column=0)

# Dropdown Box 3
label_3 = tk.Label(root, text="Country 2")
label_3.grid(row=2, column=2)
dropdown_3 = ttk.Combobox(root, values=countries)
dropdown_3.grid(row=3, column=2)

# Dropdown Box 4
label_4 = tk.Label(root, text="Competition 1 / Team 1")
label_4.grid(row=4, column=0)
dropdown_4 = ttk.Combobox(root)
dropdown_4.grid(row=5, column=0)

# Dropdown Box 5
label_5 = tk.Label(root, text="Competition 2 / Team 2")
label_5.grid(row=4, column=2)
dropdown_5 = ttk.Combobox(root)
dropdown_5.grid(row=5, column=2)

# Dropdown Box 6
label_6 = tk.Label(root, text="Team 1")
label_6.grid(row=6, column=0)
dropdown_6 = ttk.Combobox(root)
dropdown_6.grid(row=7, column=0)

# Dropdown Box 7
label_7 = tk.Label(root, text="Team 2")
label_7.grid(row=6, column=2)
dropdown_7 = ttk.Combobox(root)
dropdown_7.grid(row=7, column=2)

# Compare button
compare_button = tk.Button(root, text="Compare two teams",command=compare_two_teams_data)
compare_button.grid(row=8, column=1)

root.mainloop()
