from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import Select
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
import os
import time
import openpyxl

# Bypass SSL verification
os.environ['WDM_SSL_VERIFY'] = '0'

# Initialize the WebDriver
driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()))

# Open the target URL
url = "https://us.soccerway.com/matches/2024/07/23/korea-republic/k-league/seongnam-ilhwa/chunnam-dragons/4308862/head2head/"
driver.get(url)

# Wait for the page to load completely
driver.implicitly_wait(10)


# Function to extract options from a dropdown
def get_dropdown_options(dropdown):
    select = Select(dropdown)
    options = [option.text for option in select.options]
    return options


# Function to select an option and extract data from dependent dropdowns
def extract_data_for_selection(dropdown_index, option):
    dropdown = Select(driver.find_elements(By.CSS_SELECTOR, "select")[dropdown_index])
    dropdown.select_by_visible_text(option)
    time.sleep(2)  # Wait for the page to update

    # Re-find the dependent dropdown elements
    dependent_dropdowns = driver.find_elements(By.CSS_SELECTOR, "select")

    # Extract data from the dependent dropdowns
    dependent_data = {}
    dependent_data['Dropdown 8'] = get_dropdown_options(dependent_dropdowns[7])
    dependent_data['Dropdown 9'] = get_dropdown_options(dependent_dropdowns[8])
    dependent_data['Dropdown 11'] = get_dropdown_options(dependent_dropdowns[10])
    dependent_data['Dropdown 12'] = get_dropdown_options(dependent_dropdowns[11])

    return dependent_data


# Create a new Excel workbook and select the active sheet
wb = openpyxl.Workbook()
ws = wb.active

# Initialize row counter
row_counter = 1

# Locate the initial dropdown elements
dropdown7 = Select(driver.find_elements(By.CSS_SELECTOR, "select")[6])

dropdown7_options = [option.text for option in dropdown7.options if option.text]

# Iterate over each option in Dropdown 7 and extract data
for option in dropdown7_options:
    print(f"Selecting {option} in Dropdown 7")
    all_data = extract_data_for_selection(6, option)

    # Store the current selected country in row 1
    ws.cell(row=row_counter, column=1, value=option)

    # Move to the next row
    row_counter += 1

    dropdown8 = Select(driver.find_elements(By.CSS_SELECTOR, "select")[7])
    dropdown8_options = [option2.text for option2 in dropdown8.options if option2.text]

    for option2 in dropdown8_options:
        print(f"Selecting {option2} in Dropdown 8")
        all_data = extract_data_for_selection(7, option2)

        # Store the currently selected option2 in the next cell in the row
        ws.cell(row=row_counter, column=1, value=option2)

        # Move to the next row
        row_counter += 1

        dropdown9 = Select(driver.find_elements(By.CSS_SELECTOR, "select")[8])
        dropdown9_options = [option3.text for option3 in dropdown9.options if option3.text]

        # Store all values inside dropdown9_options in the next cells in the row
        for idx, option3 in enumerate(dropdown9_options):
            ws.cell(row=row_counter, column=idx + 1, value=option3)

        # Move to the next row
        row_counter += 1

    # Ensure to skip two rows after each main iteration
    row_counter += 3

# Save the workbook
wb.save("dropdown_options.xlsx")

# Close the WebDriver
driver.quit()
