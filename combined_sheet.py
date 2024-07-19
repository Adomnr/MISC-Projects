# import openpyxl
# import os
#
#
# def combine_workbooks(directory, output_file):
#     combined_wb = openpyxl.Workbook()
#     combined_ws = combined_wb.active
#     combined_ws.title = "Combined Data"
#
#     current_row = 1
#
#     for filename in os.listdir(directory):
#         if filename.endswith(".xlsx"):
#             filepath = os.path.join(directory, filename)
#             wb = openpyxl.load_workbook(filepath)
#             ws = wb.active
#
#             for row in ws.iter_rows(values_only=True):
#                 combined_ws.append(row)
#                 current_row += 1
#
#     combined_wb.save(output_file)
#
#
# # Specify the directory containing the workbooks and the name of the output file
# directory = "Report_Excel/"
# output_file = "combined_worked.xlsx"
#
# combine_workbooks(directory, output_file)

import openpyxl
import os

# def combine_workbooks(directory, output_file):
#     combined_wb = openpyxl.Workbook()
#     combined_ws = combined_wb.active
#     combined_ws.title = "Combined Data"
#
#     for i in range(1, 5769):
#         filename = f"Report_page_{i}.xlsx"
#         filepath = os.path.join(directory, filename)
#
#         if os.path.exists(filepath):
#             wb = openpyxl.load_workbook(filepath)
#             ws = wb.active
#
#             for row in ws.iter_rows(values_only=True):
#                 combined_ws.append(row)
#
#     combined_wb.save(output_file)
#
#
# # Specify the directory containing the workbooks and the name of the output file
# directory = "C:\\Users\\Ubaid Rehman\\Desktop\\PDF2EXCEL\\PDF2EXCEL\\New_Report"
# output_file = "series_combined_workbook.xlsx"
#
# combine_workbooks(directory, output_file)

import openpyxl
import os


def combine_workbooks(directory, output_file):
    combined_wb = openpyxl.Workbook()
    combined_ws = combined_wb.active
    combined_ws.title = "Combined Data"

    is_first_file = True

    for i in range(1, 5769):
        filename = f"Report_page{i}.xlsx"
        filepath = os.path.join(directory, filename)

        if os.path.exists(filepath):
            print(f"Processing {filename}")
            wb = openpyxl.load_workbook(filepath)
            ws = wb.active

            for row_index, row in enumerate(ws.iter_rows(values_only=True), start=1):
                # Skip the header row for subsequent files
                if not is_first_file and row_index == 1:
                    continue
                combined_ws.append(row)

            is_first_file = False
        else:
            print(f"File {filename} not found")

    combined_wb.save(output_file)
    print(f"Combined workbook saved as {output_file}")


# Specify the directory containing the workbooks and the name of the output file
directory = "C:\\Users\\Ubaid Rehman\\Desktop\\PDF2EXCEL\\PDF2EXCEL\\Report_Excel\\"
output_file = "serie_combined_workbook.xlsx"

combine_workbooks(directory, output_file)


